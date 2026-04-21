#!/usr/bin/env python3
"""Safely update the 6 point-data xlsx files with fresh pipeline values.

Reads pipeline Results CSVs produced by MATLAB processing scripts, then
uses openpyxl to overwrite specific cells in the xlsx files. openpyxl
edits cells in place and preserves all formatting (merged cells, fonts,
colours, formulae) -- unlike MATLAB's writetable/writematrix which can
corrupt the sheet layout.

Target cells in each xlsx:
  N1_*_UMi.xlsx              cols F (PL), H (OmniDS), J (OmniASA), L (OmniASD)
  U3_*_UMi.xlsx "U1" col     cols H (PL), K (DS),     N (ASA),     Q (ASD)
  N3_*_UMi.xlsx "N1" col     cols H (PL), K (DS),     N (ASA),     Q (ASD)

Usage:
    python matlab/tools/update_xlsx_from_pipeline.py
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
PD = REPO / "data" / "point_data"
BACKUP = PD / "_backup_pre_N1U1_update"


# --------------------------------------------------------------------------- #
# Pipeline Results readers
# --------------------------------------------------------------------------- #
def load_pipeline_csv(path: Path) -> list[dict]:
    with path.open(newline="") as f:
        rdr = csv.DictReader(f)
        return list(rdr)


def rx_num(s: str) -> int | None:
    """Extract RX number from names like 'R01', 'RX1', 'LOS_RX1_07-12-2024'."""
    m = re.search(r"RX?(\d+)", s)
    return int(m.group(1)) if m else None


def tx_rx_nums(s: str) -> tuple[int, int] | None:
    m = re.search(r"T(?:X)?(\d+)-R(?:X)?(\d+)", s)
    return (int(m.group(1)), int(m.group(2))) if m else None


# --------------------------------------------------------------------------- #
# Per-type updaters
# --------------------------------------------------------------------------- #
def find_header_row_end(ws) -> int:
    """Return the 1-based row where data starts. Assumes first real data row
    is the first row whose column B (TX) or column C (RX) contains a
    non-empty value matching 'TX*' or 'RX*'."""
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if (isinstance(b, str) and b.startswith("TX")) or (
            isinstance(c, str) and c.startswith("RX")
        ):
            return r
    return 3


def update_U3(xlsx_path: Path, pipe_csv: Path, pipe_cols: dict) -> int:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    pipe = load_pipeline_csv(pipe_csv)
    # Two parallel indexes:
    #   (rx_num, env)   for USC 145 where Location = "R01" etc.
    #   (distance, env) for USC 7 NLOS where Location = "NLOS_65m" (no RX no.)
    by_rx: dict[tuple[int, str], dict] = {}
    by_dist: dict[tuple[float, str], dict] = {}
    for row in pipe:
        loc = row.get("Location", "") or row.get("Location_ID", "")
        env = (row.get("Env") or row.get("Environment") or "").strip().upper()
        if not env:
            continue
        # USC 7 Location = "LOS_RX1_07-12-2024" (has RX) or "NLOS_65m" (no RX)
        if "_RX" in loc:
            import re as _re
            m = _re.search(r"_RX(\d+)", loc)
            if m:
                by_rx[(int(m.group(1)), env)] = row
        else:
            n = rx_num(loc)
            if n is not None and "_RX" not in loc and "m" not in loc.split("_")[-1]:
                by_rx[(n, env)] = row
        # Distance-based fallback index
        try:
            d = float(row.get("Distance_m", "nan"))
            if not (d != d):  # skip NaN
                by_dist[(round(d, 1), env)] = row
        except (ValueError, TypeError):
            pass

    start = find_header_row_end(ws)
    updated = 0
    for r in range(start, ws.max_row + 1):
        rx_cell = ws.cell(r, 3).value
        loc_cell = ws.cell(r, 4).value
        tr_cell = ws.cell(r, 5).value
        if not isinstance(rx_cell, str) or not isinstance(loc_cell, str):
            continue
        n = rx_num(rx_cell)
        env = loc_cell.strip().upper()

        prow = None
        # 1) try (rx, env) with env as-is
        for e in [env] + (["NLOS"] if env == "OLOS" else []):
            if n is not None and (n, e) in by_rx:
                prow = by_rx[(n, e)]
                break
        # 2) fall back to distance-based match
        if prow is None and isinstance(tr_cell, (int, float)):
            for e in [env] + (["NLOS"] if env == "OLOS" else []):
                key = (round(float(tr_cell), 1), e)
                if key in by_dist:
                    prow = by_dist[key]
                    break
        if prow is None:
            continue
        # U3 columns: 6=NYU thres, 7=USC thres, 8=USC orig (U1), etc.
        #
        # Phase-3 reference refresh (2026-04-20): we now ALSO write col 7
        # (USC thres = NYU method + USC threshold = the U3 PROPER variant per
        # paper Sec V-A "thresholds corresponding to the dataset"). USC pipeline
        # runs at USC threshold, so its `PL_NYU_dB`/`DS_NYU_ns`/`ASA_NYU_10dB`
        # /`ASD_NYU_10dB` outputs are the U3 headline values.
        #
        # Col 6 (NYU thres = NYU method + NYU threshold) requires a NYU-threshold
        # pass on USC data which the current pipeline does not run; left as
        # legacy from processing_cb_a.
        # col 8 = U1 baseline (USC method + USC threshold)
        ws.cell(r, 8).value  = round(float(prow[pipe_cols["PL"]]), 2)
        ws.cell(r, 11).value = round(float(prow[pipe_cols["DS"]]), 2)
        ws.cell(r, 14).value = round(float(prow[pipe_cols["ASA"]]), 2)
        ws.cell(r, 17).value = round(float(prow[pipe_cols["ASD"]]), 2)
        # col 7 = U3 PROPER (NYU method + USC threshold)
        if "PL_NYU" in pipe_cols and pipe_cols["PL_NYU"] in prow:
            ws.cell(r, 7).value  = round(float(prow[pipe_cols["PL_NYU"]]), 2)
            ws.cell(r, 10).value = round(float(prow[pipe_cols["DS_NYU"]]), 2)
            ws.cell(r, 13).value = round(float(prow[pipe_cols["ASA_NYU"]]), 2)
            ws.cell(r, 16).value = round(float(prow[pipe_cols["ASD_NYU"]]), 2)
        updated += 1
    wb.save(xlsx_path)
    return updated


def update_N1(xlsx_path: Path, pipe_csv: Path, pipe_cols: dict) -> int:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    pipe = load_pipeline_csv(pipe_csv)
    # Build (tx,rx) -> row dict. CSV uses TX_RX_ID = "T1-R5" format.
    pipe_by_tx_rx: dict[tuple[int, int], dict] = {}
    for row in pipe:
        key_src = row.get("TX_RX_ID") or row.get("Location") or ""
        tk = tx_rx_nums(key_src)
        if tk:
            pipe_by_tx_rx[tk] = row
    tx_cur = None
    updated = 0
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if isinstance(b, str) and b.startswith("TX"):
            mtx = re.search(r"TX(\d+)", b)
            if mtx:
                tx_cur = int(mtx.group(1))
        if not isinstance(c, str) or not c.startswith("RX") or tx_cur is None:
            continue
        m = re.match(r"RX(\d+)", c)
        if not m:
            continue
        key = (tx_cur, int(m.group(1)))
        if key not in pipe_by_tx_rx:
            continue
        prow = pipe_by_tx_rx[key]
        ws.cell(r, 6).value = round(float(prow[pipe_cols["PL"]]), 2)
        ws.cell(r, 8).value = round(float(prow[pipe_cols["DS"]]), 2)
        ws.cell(r, 10).value = round(float(prow[pipe_cols["ASA"]]), 2)
        ws.cell(r, 12).value = round(float(prow[pipe_cols["ASD"]]), 2)
        updated += 1
    wb.save(xlsx_path)
    return updated


def update_N3(xlsx_path: Path, pipe_csv: Path, pipe_cols: dict) -> int:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    pipe = load_pipeline_csv(pipe_csv)
    pipe_by_tx_rx: dict[tuple[int, int], dict] = {}
    for row in pipe:
        key_src = row.get("TX_RX_ID") or ""
        tk = tx_rx_nums(key_src)
        if tk:
            pipe_by_tx_rx[tk] = row
    tx_cur = None
    start = find_header_row_end(ws)
    updated = 0
    for r in range(start, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if isinstance(b, str) and b.startswith("TX"):
            mtx = re.search(r"TX(\d+)", b)
            if mtx:
                tx_cur = int(mtx.group(1))
        if not isinstance(c, str) or not c.startswith("RX") or tx_cur is None:
            continue
        m = re.match(r"RX(\d+)", c)
        if not m:
            continue
        key = (tx_cur, int(m.group(1)))
        if key not in pipe_by_tx_rx:
            continue
        prow = pipe_by_tx_rx[key]
        # N3 columns: 6=USC thres, 7=NYU thres, 8=NYU orig (N1), etc.
        #
        # Phase-3 reference refresh (2026-04-20): now writes col 7 (NYU thres
        # = USC method + NYU threshold = N3 PROPER variant per paper Sec V-A
        # "thresholds corresponding to the dataset"). NYU 142 pipeline runs at
        # NYU threshold, so `PL_USC_perDelayMax_dB` etc. ARE the N3 headline
        # values. NYU 7 pipeline runs both thresholds, so we ALSO write col 6.
        #
        # col 8 = N1 baseline (NYU method + NYU threshold)
        ws.cell(r, 8).value  = round(float(prow[pipe_cols["PL"]]), 2)
        ws.cell(r, 11).value = round(float(prow[pipe_cols["DS"]]), 2)
        ws.cell(r, 14).value = round(float(prow[pipe_cols["ASA"]]), 2)
        ws.cell(r, 17).value = round(float(prow[pipe_cols["ASD"]]), 2)

        # col 7 = N3 PROPER (USC method + NYU threshold)
        # NYU 142: PL_USC = 'PL_USC_perDelayMax_dB' (USC method @ NYU thresh).
        # NYU 7  : PL_USC = 'USCthr_PL_pDM_dB' which is USC method @ USC thresh
        #          (= sensitivity, col 6); we use NYU7's NYUthr_*_pDM/U cols
        #          via the dedicated keys "PL_USCmethod_NYUthr" etc.
        col7_key = "PL_USCmethod_NYUthr" if "PL_USCmethod_NYUthr" in pipe_cols else "PL_USC"
        if col7_key in pipe_cols and pipe_cols[col7_key] in prow:
            suf = "_USCmethod_NYUthr" if col7_key.endswith("USCmethod_NYUthr") else "_USC"
            ws.cell(r, 7).value  = round(float(prow[pipe_cols[f"PL{suf}"]]), 2)
            ws.cell(r, 10).value = round(float(prow[pipe_cols[f"DS{suf}"]]), 2)
            ws.cell(r, 13).value = round(float(prow[pipe_cols[f"ASA{suf}"]]), 2)
            ws.cell(r, 16).value = round(float(prow[pipe_cols[f"ASD{suf}"]]), 2)

        # col 6 = sensitivity variant (USC method + USC threshold) -- only NYU 7
        # pipeline produces this; NYU 142 leaves col 6 legacy.
        if "PL_USCmethod_USCthr" in pipe_cols and pipe_cols["PL_USCmethod_USCthr"] in prow:
            ws.cell(r, 6).value  = round(float(prow[pipe_cols["PL_USCmethod_USCthr"]]), 2)
            ws.cell(r, 9).value  = round(float(prow[pipe_cols["DS_USCmethod_USCthr"]]), 2)
            ws.cell(r, 12).value = round(float(prow[pipe_cols["ASA_USCmethod_USCthr"]]), 2)
            ws.cell(r, 15).value = round(float(prow[pipe_cols["ASD_USCmethod_USCthr"]]), 2)
        updated += 1
    wb.save(xlsx_path)
    return updated


# --------------------------------------------------------------------------- #
# Column mappings per pipeline CSV schema
# --------------------------------------------------------------------------- #
USC_COLS = {
    "PL": "PL_USC_dB",
    "DS": "DS_USC_ns",
    "ASA": "ASA_USC",
    "ASD": "ASD_USC",
    "PL_NYU":  "PL_NYU_dB",
    "DS_NYU":  "DS_NYU_ns",
    "ASA_NYU": "ASA_NYU_10dB",
    "ASD_NYU": "ASD_NYU_10dB",
}

NYU142_COLS = {
    "PL":  "PL_NYU_SUM_dB",
    "DS":  "DS_NYU_SUM_ns",
    "ASA": "ASA_NYU_10dB",
    "ASD": "ASD_NYU_10dB",
    "PL_USC":  "PL_USC_perDelayMax_dB",
    "DS_USC":  "DS_USC_perDelayMax_ns",
    "ASA_USC": "ASA_USC",
    "ASD_USC": "ASD_USC",
}

NYU7_COLS = {
    # col 8 (N1 baseline) = NYU method + NYU threshold
    "PL":  "NYUthr_PL_SUM_dB",
    "DS":  "NYUthr_DS_SUM_ns",
    "ASA": "NYUthr_ASA_N10",
    "ASD": "NYUthr_ASD_N10",
    # col 7 (N3 proper, "NYU thres") = USC method + NYU threshold
    "PL_USCmethod_NYUthr":  "NYUthr_PL_pDM_dB",
    "DS_USCmethod_NYUthr":  "NYUthr_DS_pDM_ns",
    "ASA_USCmethod_NYUthr": "NYUthr_ASA_U",
    "ASD_USCmethod_NYUthr": "NYUthr_ASD_U",
    # col 6 (sensitivity, "USC thres") = USC method + USC threshold
    "PL_USCmethod_USCthr":  "USCthr_PL_pDM_dB",
    "DS_USCmethod_USCthr":  "USCthr_DS_pDM_ns",
    "ASA_USCmethod_USCthr": "USCthr_ASA_U",
    "ASD_USCmethod_USCthr": "USCthr_ASD_U",
    # back-compat alias
    "PL_USC":  "USCthr_PL_pDM_dB",
    "DS_USC":  "USCthr_DS_pDM_ns",
    "ASA_USC": "USCthr_ASA_U",
    "ASD_USC": "USCthr_ASD_U",
}


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    print("\nUpdating xlsx with pipeline values (openpyxl, in-place)...\n")

    usc145_csv = REPO / "matlab/processing/usc_145/Results/USC145GHz_Full_Results.csv"
    usc7_csv = REPO / "matlab/processing/usc_7/Results/USC7GHz_NewData_Results.csv"
    nyu142_csv = REPO / "matlab/processing/nyu_142/Results/NYU142GHz_Method_Comparison_Results.csv"
    nyu7_csv = REPO / "matlab/processing/nyu_7/Results/NYU7GHz_Method_Comparison_Results.csv"

    tasks = [
        ("U3", PD / "U3_142_UMi.xlsx", usc145_csv, USC_COLS),
        ("U3", PD / "U3_7_UMi.xlsx",   usc7_csv,   USC_COLS),
        ("N1", PD / "N1_142_UMi.xlsx", nyu142_csv, NYU142_COLS),
        ("N1", PD / "N1_7_UMi.xlsx",   nyu7_csv,   NYU7_COLS),
        ("N3", PD / "N3_142_UMi.xlsx", nyu142_csv, NYU142_COLS),
        ("N3", PD / "N3_7_UMi.xlsx",   nyu7_csv,   NYU7_COLS),
    ]

    total = 0
    for kind, xlsx, csv_p, cols in tasks:
        if not xlsx.exists():
            print(f"  [skip] {xlsx.name} (missing)")
            continue
        if not csv_p.exists():
            print(f"  [skip] {xlsx.name} (pipeline CSV missing: {csv_p.name})")
            continue
        fn = {"U3": update_U3, "N1": update_N1, "N3": update_N3}[kind]
        n = fn(xlsx, csv_p, cols)
        print(f"  [{kind}] {xlsx.name:<24s} {n:3d} links updated")
        total += n

    print(f"\nTotal: {total} links updated across {len(tasks)} xlsx files.")


if __name__ == "__main__":
    main()
