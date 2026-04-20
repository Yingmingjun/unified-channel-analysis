#!/usr/bin/env python3
"""Extract U1 values from U3 xlsx 'USC orig (U1)' column into standalone
CSVs matching the N1 table layout, for inclusion in the supplement.

Output files (written to figures/matlab/):
    table_U1_145.csv
    table_U1_7.csv

Columns match the N1_*_UMi.xlsx layout:
    Freq. | TX | RX | Loc Type | TR Sep | PL | Mean Dir DS | Omni DS |
    Mean Lobe ASA | Omni ASA | Mean Lobe ASD | Omni ASD |
    Mean Lobe ZSA | Omni ZSA | Mean Lobe ZSD | Omni ZSD

We populate PL/DS/ASA/ASD from U3 xlsx col H/K/N/Q. 'Mean Dir' / 'Mean Lobe'
and ZSA/ZSD columns are left empty (those aren't computed in the current
pipeline).
"""
from __future__ import annotations

import csv
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
PD = REPO / "data" / "point_data"
OUT = REPO / "figures" / "matlab"
OUT.mkdir(parents=True, exist_ok=True)


def extract(u3_xlsx: Path, out_csv: Path, freq_label: str) -> int:
    wb = load_workbook(u3_xlsx)
    ws = wb.active

    # Header rows = 2 or 3 depending on xlsx; skip until we find a real TX/RX
    data_start = 3
    for r in range(1, min(6, ws.max_row + 1)):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.startswith("TX"):
            data_start = r
            break

    rows = [[
        "Freq.", "TX", "RX", "Loc Type", "TR Sep",
        "PL", "Mean Dir DS", "Omni DS",
        "Mean Lobe ASA", "Omni ASA",
        "Mean Lobe ASD", "Omni ASD",
        "Mean Lobe ZSA", "Omni ZSA",
        "Mean Lobe ZSD", "Omni ZSD",
    ]]

    tx_cur = ""
    count = 0
    for r in range(data_start, ws.max_row + 1):
        tx = ws.cell(r, 2).value
        rx = ws.cell(r, 3).value
        loc = ws.cell(r, 4).value
        tr = ws.cell(r, 5).value
        pl = ws.cell(r, 8).value    # H = USC orig PL
        ds = ws.cell(r, 11).value   # K = USC orig DS
        asa = ws.cell(r, 14).value  # N = USC orig ASA
        asd = ws.cell(r, 17).value  # Q = USC orig ASD

        if isinstance(tx, str) and tx.startswith("TX"):
            tx_cur = tx
        import re as _re
        if not isinstance(rx, str) or not _re.match(r"RX\d+$", rx):
            continue

        rows.append([
            freq_label,
            tx_cur if count == 0 else "",
            rx,
            loc or "",
            tr if tr is not None else "",
            pl if pl is not None else "",
            "",   # Mean Dir DS (not computed)
            ds if ds is not None else "",
            "",   # Mean Lobe ASA
            asa if asa is not None else "",
            "",   # Mean Lobe ASD
            asd if asd is not None else "",
            "", "", "", "",  # ZSA / ZSD not computed
        ])
        count += 1

    with out_csv.open("w", newline="") as f:
        csv.writer(f).writerows(rows)
    return count


def main() -> None:
    n1 = extract(PD / "U3_142_UMi.xlsx", OUT / "table_U1_145.csv", "145 GHz")
    n2 = extract(PD / "U3_7_UMi.xlsx",   OUT / "table_U1_7.csv",   "6.75 GHz")
    print(f"[extract_U1] wrote {OUT / 'table_U1_145.csv'} ({n1} links)")
    print(f"[extract_U1] wrote {OUT / 'table_U1_7.csv'}   ({n2} links)")


if __name__ == "__main__":
    main()
